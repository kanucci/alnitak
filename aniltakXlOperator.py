from openpyxl import load_workbook

class oPlan():

    def __init__(self, address, data_list):
        self.address = address
        self.workbook = load_workbook(self.address)
        self.data_list = data_list
        #self.sheet_list = self.workbook.sheetnames


    def operator(self):
        wb = self.workbook
        sheet_list = wb.sheetnames
        employee_register = []
        er_list = self.data_list

        for sheet_name in sheet_list:
            oSheet = wb[sheet_name]
            for rSheet in oSheet:

                if rSheet[3].value == 'MATRICULA':
                    continue

                if rSheet[5].value == 'Retorno ao trabalho' or rSheet[5].value == ' Retorno ao trabalho':
                    employee_register = [rSheet[0].value, rSheet[3].value, rSheet[4].value, 'RetornoTrabalho']
                    er_list.append(employee_register)

                elif rSheet[5].value == 'Periódico':
                    employee_register = [rSheet[0].value, rSheet[3].value, rSheet[4].value, 'Periodico']
                    er_list.append(employee_register)

                else:

                    employee_register = [rSheet[0].value, rSheet[3].value, rSheet[4].value, rSheet[5].value]
                    er_list.append(employee_register)

        wb.close()

        return er_list

        def updateXl(self):
            wb = self.workbook
            she
